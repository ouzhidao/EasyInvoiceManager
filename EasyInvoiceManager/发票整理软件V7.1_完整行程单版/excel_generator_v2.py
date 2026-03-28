"""
Excel生成模块 V2.0
支持多子表：
1. 发票汇总（增值税发票）
2. 铁路电子客票
3. 识别失败
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelGeneratorV2:
    """Excel生成器 V2.0 - 支持多子表"""
    
    # 边框样式
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self):
        self.wb = None
    
    def generate(self, vat_invoices: list, railway_invoices: list, itinerary_invoices: list, failed_invoices: list, output_path: str):
        """
        生成Excel文件
        
        Args:
            vat_invoices: 增值税发票列表
            railway_invoices: 铁路电子客票列表
            failed_invoices: 识别失败列表
            output_path: 输出路径
            
        Returns:
            (文件路径, 总金额, 类型汇总)
        """
        self.wb = Workbook()
        
        # 移除默认的sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        # 创建三个子表
        self._create_vat_sheet(vat_invoices)
        self._create_railway_sheet(railway_invoices)
        self._create_itinerary_sheet(itinerary_invoices)
        self._create_failed_sheet(failed_invoices)
        
        # 计算统计信息
        total_amount = self._calculate_total_amount(vat_invoices, railway_invoices, itinerary_invoices)
        type_summary = self._calculate_type_summary(vat_invoices, railway_invoices, itinerary_invoices)
        
        # 保存文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        output_file = Path(output_path) / f"发票清单_{timestamp}.xlsx"
        self.wb.save(output_file)
        
        return str(output_file), total_amount, type_summary
    
    def _create_vat_sheet(self, invoices: list):
        """创建增值税发票子表"""
        ws = self.wb.create_sheet("增值税发票", 0)
        
        # 剔除重复发票
        non_duplicate = [inv for inv in invoices if not inv.get('is_duplicate', False)]
        
        # 表头
        headers = ['序号', '新文件名', '发票号码', '购买方名称', '内容', '金额', '开票方', '日期', '类型', '识别方式']
        self._write_headers(ws, headers)
        
        # 数据
        for idx, inv in enumerate(non_duplicate, 1):
            ws.append([
                idx,
                inv.get('new_filename', ''),
                inv.get('invoice_num_full', ''),
                inv.get('buyer_name', ''),
                inv.get('content', ''),
                inv.get('amount', 0),
                inv.get('seller_name', ''),
                inv.get('date', ''),
                inv.get('invoice_type', ''),
                inv.get('recognize_method', 'API')
            ])
        
        # 设置列宽
        col_widths = [8, 45, 20, 35, 20, 12, 35, 12, 18, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 设置样式
        self._apply_data_styles(ws, len(non_duplicate) + 1, len(headers))
        
        # 添加汇总区域
        self._add_vat_summary(ws, invoices, non_duplicate)
        
        # 设置打印选项
        self._setup_print_options(ws, 'A', len(non_duplicate) + 10)
    
    def _create_railway_sheet(self, invoices: list):
        """创建铁路电子客票子表"""
        ws = self.wb.create_sheet("铁路电子客票", 1)
        
        # 剔除重复发票
        non_duplicate = [inv for inv in invoices if not inv.get('is_duplicate', False)]
        
        # 表头（14列）
        headers = [
            '序号', '新文件名', '发票号码', '购买方名称', '乘车人',
            '出发地', '到达地', '乘车日期', '开车时间', '席别',
            '金额', '开票日期', '识别方式', '备注'
        ]
        self._write_headers(ws, headers)
        
        # 数据
        for idx, inv in enumerate(non_duplicate, 1):
            ws.append([
                idx,
                inv.get('new_filename', ''),
                inv.get('invoice_num', ''),
                inv.get('buyer_name', ''),
                inv.get('passenger_name', ''),
                inv.get('departure', ''),
                inv.get('arrival', ''),
                inv.get('travel_date', ''),
                inv.get('departure_time', ''),
                inv.get('seat_type', ''),
                inv.get('amount', 0),
                inv.get('invoice_date', ''),
                inv.get('recognize_method', 'PaddleOCR'),
                ''  # 备注
            ])
        
        # 设置列宽
        col_widths = [8, 50, 22, 30, 10, 12, 12, 12, 10, 15, 10, 12, 12, 15]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 设置样式
        self._apply_data_styles(ws, len(non_duplicate) + 1, len(headers))
        
        # 添加汇总区域
        self._add_railway_summary(ws, invoices, non_duplicate)
        
        # 设置打印选项
        summary_row = len(non_duplicate) + 8
        self._setup_print_options(ws, 'A', summary_row)
    
    def _create_failed_sheet(self, failed_list: list):
        """创建识别失败子表"""
        ws = self.wb.create_sheet("识别失败", 2)
        
        # 表头
        headers = ['序号', '原文件名', '失败原因', '文件路径']
        self._write_headers(ws, headers)
        
        # 数据
        for idx, item in enumerate(failed_list, 1):
            ws.append([
                idx,
                item.get('filename', ''),
                item.get('reason', ''),
                item.get('file_path', '')
            ])
        
        # 设置列宽
        col_widths = [8, 40, 40, 60]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 设置样式
        self._apply_data_styles(ws, len(failed_list) + 1, len(headers))
        
        # 设置打印选项
        self._setup_print_options(ws, 'A', max(len(failed_list) + 1, 10))
    
    def _write_headers(self, ws, headers: list):
        """写入表头"""
        ws.append(headers)
        
        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        ws.row_dimensions[1].height = 25
    
    def _apply_data_styles(self, ws, last_row: int, last_col: int):
        """应用数据区域样式"""
        for row in range(2, last_row + 1):
            for col in range(1, last_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(size=10)
            ws.row_dimensions[row].height = 20
    
    def _add_vat_summary(self, ws, all_invoices: list, non_duplicate: list):
        """添加增值税发票汇总"""
        start_row = len(non_duplicate) + 3
        
        # 按类型汇总
        type_summary = {}
        for inv in non_duplicate:
            inv_type = inv.get('invoice_type', '其他')
            if inv_type not in type_summary:
                type_summary[inv_type] = {'count': 0, 'amount': 0.0}
            type_summary[inv_type]['count'] += 1
            type_summary[inv_type]['amount'] += inv.get('amount', 0)
        
        # 汇总标题
        ws.merge_cells(f'A{start_row}:B{start_row}')
        title_cell = ws.cell(row=start_row, column=1, value="汇总统计")
        title_cell.font = Font(bold=True, size=12, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2E75B6", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[start_row].height = 25
        
        # 类型汇总数据
        current_row = start_row + 1
        for inv_type, data in sorted(type_summary.items()):
            ws.cell(row=current_row, column=1, value=inv_type)
            ws.cell(row=current_row, column=2, value=f"{data['count']} 张")
            ws.cell(row=current_row, column=3, value=f"{data['amount']:.2f} 元")
            
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col)
                cell.font = Font(size=11)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            ws.row_dimensions[current_row].height = 22
            current_row += 1
        
        # 合计行
        total_count = len(non_duplicate)
        total_amount = sum(r.get('amount', 0) for r in non_duplicate)
        
        ws.cell(row=current_row, column=1, value="合计")
        ws.cell(row=current_row, column=2, value=f"{total_count} 张")
        ws.cell(row=current_row, column=3, value=f"{total_amount:.2f} 元")
        
        for col in range(1, 4):
            cell = ws.cell(row=current_row, column=col)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="C65911", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[current_row].height = 25
    
    def _add_railway_summary(self, ws, all_invoices: list, non_duplicate: list):
        """添加铁路客票汇总"""
        start_row = len(non_duplicate) + 3
        
        # 统计信息
        total_count = len(non_duplicate)
        total_amount = sum(r.get('amount', 0) for r in non_duplicate)
        
        # 汇总标题
        ws.merge_cells(f'A{start_row}:C{start_row}')
        title_cell = ws.cell(row=start_row, column=1, value="铁路电子客票汇总")
        title_cell.font = Font(bold=True, size=12, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2E75B6", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[start_row].height = 25
        
        # 统计数据
        summary_data = [
            ['总张数', f'{total_count} 张', ''],
            ['总金额', f'{total_amount:.2f} 元', ''],
        ]
        
        current_row = start_row + 1
        for row_data in summary_data:
            ws.cell(row=current_row, column=1, value=row_data[0])
            ws.cell(row=current_row, column=2, value=row_data[1])
            
            for col in range(1, 3):
                cell = ws.cell(row=current_row, column=col)
                cell.font = Font(size=11)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            ws.row_dimensions[current_row].height = 22
            current_row += 1
        
        # 合计行样式
        ws.cell(row=current_row, column=1, value="合计")
        ws.cell(row=current_row, column=2, value=f"{total_count} 张")
        ws.cell(row=current_row, column=3, value=f"{total_amount:.2f} 元")
        
        for col in range(1, 4):
            cell = ws.cell(row=current_row, column=col)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="C65911", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[current_row].height = 25
    
    def _setup_print_options(self, ws, start_col: str, end_row: int):
        """设置打印选项"""
        # 设置打印区域
        last_col = get_column_letter(ws.max_column)
        ws.print_area = f'{start_col}1:{last_col}{end_row}'
        
        # 横向打印（A4纸）
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        # 每页重复表头
        ws.print_title_rows = '1:1'
        
        # 页边距（英寸）
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        
        # 居中打印
        ws.print_options.horizontalCentered = True
        
        # 页眉页脚
        ws.oddHeader.center.text = ws.title
        ws.oddHeader.center.size = 14
        ws.oddHeader.center.font = "Arial,Bold"
        ws.oddFooter.center.text = "第 &P 页，共 &N 页"
        ws.oddFooter.center.size = 10
    
    def _calculate_total_amount(self, vat_invoices: list, railway_invoices: list, itinerary_invoices: list = None) -> float:
        """计算总金额（不包含重复发票）"""
        vat_amount = sum(
            inv.get('amount', 0) 
            for inv in vat_invoices 
            if not inv.get('is_duplicate', False)
        )
        railway_amount = sum(
            inv.get('amount', 0) 
            for inv in railway_invoices 
            if not inv.get('is_duplicate', False)
        )
        itinerary_amount = sum(
            inv.get('amount', 0) 
            for inv in (itinerary_invoices or [])
            if not inv.get('is_duplicate', False)
        )
        return round(vat_amount + railway_amount + itinerary_amount, 2)
    
    def _calculate_type_summary(self, vat_invoices: list, railway_invoices: list, itinerary_invoices: list = None) -> dict:
        """计算按类型汇总"""
        summary = {}
        
        # 增值税发票汇总
        for inv in vat_invoices:
            if inv.get('is_duplicate', False):
                continue
            inv_type = inv.get('invoice_type', '其他发票')
            if inv_type not in summary:
                summary[inv_type] = {'count': 0, 'amount': 0.0}
            summary[inv_type]['count'] += 1
            summary[inv_type]['amount'] += inv.get('amount', 0)
        
        # 铁路客票汇总
        railway_count = sum(
            1 for inv in railway_invoices 
            if not inv.get('is_duplicate', False)
        )
        railway_amount = sum(
            inv.get('amount', 0) 
            for inv in railway_invoices 
            if not inv.get('is_duplicate', False)
        )
        
        if railway_count > 0:
            summary['铁路电子客票'] = {
                'count': railway_count,
                'amount': round(railway_amount, 2)
            }
        
        # 行程单 statistics
        if itinerary_invoices:
            itinerary_count = sum(
                1 for inv in itinerary_invoices 
                if not inv.get("is_duplicate", False)
            )
            itinerary_amount = sum(
                inv.get("amount", 0) 
                for inv in itinerary_invoices 
                if not inv.get("is_duplicate", False)
            )
            
            if itinerary_count > 0:
                summary["行程单"] = {
                    "count": itinerary_count,
                    "amount": round(itinerary_amount, 2)
                }
        
        return summary


    def _create_itinerary_sheet(self, invoices: list):
        """创建行程单子表 - 简化版"""
        ws = self.wb.create_sheet("行程单", 2)
        
        # 剔除重复
        non_duplicate = [inv for inv in invoices if not inv.get('is_duplicate', False)]
        
        # 表头（简化）
        headers = ['序号', '文件名', '类型', '金额']
        self._write_headers(ws, headers)
        
        # 数据
        for idx, inv in enumerate(non_duplicate, 1):
            ws.append([
                idx,
                inv.get('new_filename', ''),
                inv.get('sub_type', '其他'),
                inv.get('amount', 0)
            ])
        
        # 调整列宽
        col_widths = [8, 50, 12, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 应用样式
        self._apply_data_styles(ws, len(non_duplicate) + 1, len(headers))
        
        # 添加汇总
        self._add_itinerary_summary(ws, invoices, non_duplicate)
        
        # 设置打印选项
        summary_row = len(non_duplicate) + 8
        self._setup_print_options(ws, 'A', summary_row)
    def _add_itinerary_summary(self, ws, all_invoices: list, non_duplicate: list):
        """��ӵ�̵���ۣ��������֣���"""
        start_row = len(non_duplicate) + 3
        
        # ����ͳ��
        type_summary = {}
        for inv in non_duplicate:
            sub_type = inv.get('sub_type', '火车')
            if sub_type not in type_summary:
                type_summary[sub_type] = {'count': 0, 'amount': 0.0}
            type_summary[sub_type]['count'] += 1
            type_summary[sub_type]['amount'] += inv.get('amount', 0)
        
        # ���ܱ���
        ws.merge_cells(f'A{start_row}:D{start_row}')
        title_cell = ws.cell(row=start_row, column=1, value="行程单汇总")
        title_cell.font = Font(bold=True, size=12, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2E75B6", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[start_row].height = 25
        
        # ����ͳ��
        current_row = start_row + 1
        for sub_type in sorted(type_summary.keys()):
            data = type_summary[sub_type]
            ws.cell(row=current_row, column=1, value=sub_type)
            ws.cell(row=current_row, column=2, value=f"{data['count']} ��")
            ws.cell(row=current_row, column=3, value=f"{data['amount']:.2f} Ԫ")
            
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col)
                cell.font = Font(size=11)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            ws.row_dimensions[current_row].height = 22
            current_row += 1
        
        # �ϼ�
        total_count = len(non_duplicate)
        total_amount = sum(r.get('amount', 0) for r in non_duplicate)
        
        ws.cell(row=current_row, column=1, value="�ϼ�")
        ws.cell(row=current_row, column=2, value=f"{total_count} ��")
        ws.cell(row=current_row, column=3, value=f"{total_amount:.2f} Ԫ")
        
        for col in range(1, 4):
            cell = ws.cell(row=current_row, column=col)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="C65911", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[current_row].height = 25
