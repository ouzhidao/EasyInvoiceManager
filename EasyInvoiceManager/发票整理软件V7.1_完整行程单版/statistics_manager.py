"""
统计管理模块
"""
import json
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
from dataclasses import dataclass, asdict


@dataclass
class InvoiceRecord:
    """单张发票记录"""
    task_id: str
    file_name: str
    file_path: str
    invoice_num: str
    amount: float
    invoice_type: str
    invoice_date: str  # YYYY-MM
    seller_name: str
    buyer_name: str  # 购买方名称
    is_success: bool
    is_duplicate: bool
    error_msg: str
    process_time: str
    output_path: str


class StatisticsManager:
    """统计管理器"""
    
    def __init__(self):
        self.data_dir = Path.home() / '.invoice_processor' / 'statistics'
        self.data_dir.mkdir(parents=True, exist_ok=True)
        
        self.db_path = self.data_dir / 'invoice_statistics.db'
        self._init_database()
        
        self.current_task_id: Optional[str] = None
        self.current_records: List[InvoiceRecord] = []
    
    def _init_database(self):
        """初始化数据库"""
        conn = sqlite3.connect(str(self.db_path))
        cursor = conn.cursor()
        
        # 创建发票记录表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS invoice_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_id TEXT,
                file_name TEXT,
                file_path TEXT,
                invoice_num TEXT,
                amount REAL,
                invoice_type TEXT,
                invoice_date TEXT,
                seller_name TEXT,
                buyer_name TEXT,
                is_success INTEGER,
                is_duplicate INTEGER,
                error_msg TEXT,
                process_time TEXT,
                output_path TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # 检查并添加 buyer_name 列（兼容旧版本数据库）
        try:
            cursor.execute('SELECT buyer_name FROM invoice_records LIMIT 1')
        except sqlite3.OperationalError:
            # 列不存在，添加它
            try:
                cursor.execute('ALTER TABLE invoice_records ADD COLUMN buyer_name TEXT')
                print("数据库升级：已添加 buyer_name 列")
            except sqlite3.OperationalError as e:
                print(f"添加 buyer_name 列失败（可能已存在）: {e}")
        
        # 创建任务汇总表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS task_summary (
                task_id TEXT PRIMARY KEY,
                total_count INTEGER,
                success_count INTEGER,
                failed_count INTEGER,
                duplicate_count INTEGER,
                total_amount REAL,
                output_folder TEXT,
                process_date TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def start_task(self) -> str:
        """开始新任务"""
        self.current_task_id = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.current_records = []
        return self.current_task_id
    
    def add_record(self, record: InvoiceRecord):
        """添加记录"""
        self.current_records.append(record)
        
        # 实时保存到数据库
        conn = sqlite3.connect(str(self.db_path))
        cursor = conn.cursor()
        
        # 确保 buyer_name 不为 None
        buyer_name = record.buyer_name if record.buyer_name else ''
        
        cursor.execute('''
            INSERT INTO invoice_records 
            (task_id, file_name, file_path, invoice_num, amount, invoice_type, 
             invoice_date, seller_name, buyer_name, is_success, is_duplicate, error_msg, 
             process_time, output_path)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            record.task_id, record.file_name, record.file_path, record.invoice_num,
            record.amount, record.invoice_type, record.invoice_date, record.seller_name,
            buyer_name,  # 添加购买方名称（确保不为None）
            1 if record.is_success else 0, 1 if record.is_duplicate else 0,
            record.error_msg, record.process_time, record.output_path
        ))
        
        conn.commit()
        conn.close()
    
    def save_task_summary(self, total: int, success: int, failed: int, 
                         duplicate: int, total_amount: float, output_folder: str):
        """保存任务汇总"""
        conn = sqlite3.connect(str(self.db_path))
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT OR REPLACE INTO task_summary 
            (task_id, total_count, success_count, failed_count, duplicate_count,
             total_amount, output_folder, process_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            self.current_task_id, total, success, failed, duplicate,
            total_amount, output_folder, datetime.now().strftime('%Y-%m-%d')
        ))
        
        conn.commit()
        conn.close()
    
    def get_current_statistics(self) -> Dict:
        """获取当前任务统计"""
        total = len(self.current_records)
        success = sum(1 for r in self.current_records if r.is_success)
        failed = total - success
        duplicate = sum(1 for r in self.current_records if r.is_duplicate and r.is_success)
        
        # 按类型统计
        type_stats = {}
        for r in self.current_records:
            if r.is_success:
                type_name = r.invoice_type or '其他'
                if type_name not in type_stats:
                    type_stats[type_name] = {'count': 0, 'amount': 0}
                type_stats[type_name]['count'] += 1
                type_stats[type_name]['amount'] += r.amount
        
        # 按年月统计
        month_stats = {}
        for r in self.current_records:
            if r.is_success and r.invoice_date:
                # invoice_date 格式: YYYY-MM
                month = r.invoice_date[:7] if len(r.invoice_date) >= 7 else r.invoice_date
                if month not in month_stats:
                    month_stats[month] = {'count': 0, 'amount': 0}
                month_stats[month]['count'] += 1
                month_stats[month]['amount'] += r.amount
        
        return {
            'task_id': self.current_task_id,
            'total': total,
            'success': success,
            'failed': failed,
            'duplicate': duplicate,
            'type_statistics': type_stats,
            'month_statistics': month_stats
        }
    
    def get_history_tasks(self, limit: int = 50) -> List[Dict]:
        """获取历史任务"""
        conn = sqlite3.connect(str(self.db_path))
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT * FROM task_summary 
            ORDER BY created_at DESC 
            LIMIT ?
        ''', (limit,))
        
        columns = [description[0] for description in cursor.description]
        results = []
        
        for row in cursor.fetchall():
            results.append(dict(zip(columns, row)))
        
        conn.close()
        return results
    
    def export_to_excel(self, output_path: str, task_id: str = None) -> str:
        """导出统计到Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        
        # 确定要导出的任务
        export_task_id = task_id or self.current_task_id
        
        # 1. 汇总表
        ws_summary = wb.active
        ws_summary.title = "汇总统计"
        
        # 获取数据
        conn = sqlite3.connect(str(self.db_path))
        cursor = conn.cursor()
        
        # 任务基本信息
        cursor.execute('''
            SELECT * FROM task_summary WHERE task_id = ?
        ''', (export_task_id,))
        
        task_info = cursor.fetchone()
        if task_info:
            ws_summary.append(['任务ID', task_info[0]])
            ws_summary.append(['处理日期', task_info[7]])
            ws_summary.append(['总数量', task_info[1]])
            ws_summary.append(['成功', task_info[2]])
            ws_summary.append(['失败', task_info[3]])
            ws_summary.append(['重复', task_info[4]])
            ws_summary.append(['总金额', task_info[5]])
            ws_summary.append(['输出文件夹', task_info[6]])
        
        # 2. 按类型统计
        ws_type = wb.create_sheet("按类型统计")
        ws_type.append(['发票类型', '数量', '金额'])
        
        cursor.execute('''
            SELECT invoice_type, COUNT(*), SUM(amount) 
            FROM invoice_records 
            WHERE task_id = ? AND is_success = 1
            GROUP BY invoice_type
        ''', (export_task_id,))
        
        for row in cursor.fetchall():
            ws_type.append(list(row))
        
        # 3. 按年月统计
        ws_month = wb.create_sheet("按年月统计")
        ws_month.append(['年月', '数量', '金额'])
        
        cursor.execute('''
            SELECT invoice_date, COUNT(*), SUM(amount) 
            FROM invoice_records 
            WHERE task_id = ? AND is_success = 1
            GROUP BY invoice_date
            ORDER BY invoice_date
        ''', (export_task_id,))
        
        for row in cursor.fetchall():
            ws_month.append(list(row))
        
        # 4. 明细表
        ws_detail = wb.create_sheet("明细")
        ws_detail.append(['文件名', '发票号码', '金额', '类型', '日期', '销售方', '购买方', '状态', '是否重复'])
        
        cursor.execute('''
            SELECT file_name, invoice_num, amount, invoice_type, invoice_date, 
                   seller_name, buyer_name, is_success, is_duplicate
            FROM invoice_records 
            WHERE task_id = ?
            ORDER BY id
        ''', (export_task_id,))
        
        for row in cursor.fetchall():
            status = '成功' if row[7] else '失败'
            is_dup = '是' if row[8] else '否'
            ws_detail.append([row[0], row[1], row[2], row[3], row[4], row[5], row[6], status, is_dup])
        
        conn.close()
        
        # 设置样式
        for ws in [ws_type, ws_month, ws_detail]:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", fill_type="solid")
        
        # 保存
        output_file = Path(output_path) / f"统计报表_{export_task_id}.xlsx"
        counter = 1
        while output_file.exists():
            output_file = Path(output_path) / f"统计报表_{export_task_id}_{counter}.xlsx"
            counter += 1
        
        wb.save(output_file)
        return str(output_file)
