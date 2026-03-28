"""
Excel生成模块 - 优化版（剔除重复、横向打印、汇总）
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    def generate(self, invoices: list, output_path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "发票汇总"
        
        # 剔除重复发票
        non_duplicate_invoices = [inv for inv in invoices if not inv.get('is_duplicate', False)]
        
        # 按类型分组统计（用于汇总）
        type_summary = {}
        for inv in non_duplicate_invoices:
            inv_type = inv.get('invoice_type', '其他')
            if inv_type not in type_summary:
                type_summary[inv_type] = {'count': 0, 'amount': 0.0}
            type_summary[inv_type]['count'] += 1
            type_summary[inv_type]['amount'] += inv.get('amount', 0)
        
        # 计算总金额
        total_amount = sum(inv.get('amount', 0) for inv in non_duplicate_invoices)
        total_count = len(non_duplicate_invoices)
        
        # 标题 - 添加购买方名称列，内容列显示两个星号之间的原始内容
        headers = ['序号', '新文件名', '发票号码', '购买方名称', '内容', '金额', '开票方', '日期', '类型']
        ws.append(headers)
        
        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 设置表头行高
        ws.row_dimensions[1].height = 25
        
        # 数据（只写入非重复发票）
        for idx, inv in enumerate(non_duplicate_invoices, 1):
            ws.append([
                idx,
                inv.get('new_filename', ''),
                inv.get('invoice_num_full', ''),
                inv.get('buyer_name', ''),  # 购买方名称
                inv.get('content', ''),     # 内容（两个星号之间的原始内容）
                inv.get('amount', 0),
                inv.get('seller_name', ''),
                inv.get('date', ''),
                inv.get('invoice_type', '')
            ])
        
        # 添加汇总区域（空一行后）
        summary_start_row = total_count + 3
        
        # 汇总标题
        ws.merge_cells(f'A{summary_start_row}:B{summary_start_row}')
        summary_title = ws.cell(row=summary_start_row, column=1, value="汇总统计")
        summary_title.font = Font(bold=True, size=12, color="FFFFFF")
        summary_title.fill = PatternFill(start_color="2E75B6", fill_type="solid")
        summary_title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[summary_start_row].height = 25
        
        # 按类型汇总
        current_row = summary_start_row + 1
        for inv_type, data in sorted(type_summary.items()):
            ws.cell(row=current_row, column=1, value=inv_type)
            ws.cell(row=current_row, column=2, value=f"{data['count']} 张")
            ws.cell(row=current_row, column=3, value=f"{data['amount']:.2f} 元")
            
            # 样式
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col)
                cell.font = Font(size=11)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            ws.row_dimensions[current_row].height = 22
            current_row += 1
        
        # 合计行
        ws.cell(row=current_row, column=1, value="合计")
        ws.cell(row=current_row, column=2, value=f"{total_count} 张")
        ws.cell(row=current_row, column=3, value=f"{total_amount:.2f} 元")
        
        for col in range(1, 4):
            cell = ws.cell(row=current_row, column=col)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="C65911", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[current_row].height = 25
        
        # 列宽设置
        ws.column_dimensions['A'].width = 8   # 序号
        ws.column_dimensions['B'].width = 45  # 新文件名
        ws.column_dimensions['C'].width = 20  # 发票号码
        ws.column_dimensions['D'].width = 35  # 购买方名称
        ws.column_dimensions['E'].width = 20  # 内容（两个星号之间的原始内容）
        ws.column_dimensions['F'].width = 12  # 金额
        ws.column_dimensions['G'].width = 35  # 开票方
        ws.column_dimensions['H'].width = 12  # 日期
        ws.column_dimensions['I'].width = 18  # 类型
        
        # 设置数据区域样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 注意：数据行从第2行开始，到 total_count + 1 行结束（因为第1行是表头）
        for row in range(2, total_count + 2):
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(size=10)
            ws.row_dimensions[row].height = 20
        
        # 金额列右对齐（现在金额在第6列）
        for row in range(2, total_count + 2):
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='right', vertical='center')
        
        # 设置打印区域
        ws.print_area = f'A1:I{current_row}'
        
        # 设置横向打印（A4纸）
        ws.page_setup.orientation = 'landscape'  # 横向
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        # 设置打印标题（每页重复表头）
        ws.print_title_rows = '1:1'
        
        # 设置页边距（单位：英寸）
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3
        
        # 设置居中打印
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = False
        
        # 页眉页脚
        ws.oddHeader.center.text = "发票清单"
        ws.oddHeader.center.size = 14
        ws.oddHeader.center.font = "Arial,Bold"
        
        ws.oddFooter.center.text = "第 &P 页，共 &N 页"
        ws.oddFooter.center.size = 10
        
        # 保存
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        output_file = Path(output_path) / f"发票清单_{timestamp}.xlsx"
        wb.save(output_file)
        
        return str(output_file), total_amount, type_summary
